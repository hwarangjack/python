import openpyxl
import os
import pandas as pd
import shutil
import re
import openpyxl


# 대상파일 획정하기
path='C:\\users\\jjjac\\desktop\\HR\\☆ 자문'
how='2022년\s*임금대장\s*.+[.]x+'

empty_list=[]

def checkfile(path, target):

    cp=re.compile(target)

    for i in os.listdir(path):
        abs_path = os.path.join(path,i)
        if os.path.isfile(abs_path):    #os.path.join 함수는 절대경로가 있는 파일만 Read 할 수 있는 것 같음
            if cp.match(i):
                empty_list.append(abs_path)
        
        elif os.path.isdir(abs_path):
            checkfile(abs_path, target)

checkfile(path,how)

# 대상파일을 현재작업폴더에 복사하기
for i in empty_list:

    print(i)
    wb = openpyxl.load_workbook(i)
    ws = wb['(조건)']
    ws['p5'] = 0.03495
    ws['p6'] = 0.1227

    ws['o17'] = "2022.1~2022.6"
    ws['p17'] = 0.045
    ws['q17'] = 0.03495
    ws['r17'] = 0.1227
    ws['s17'] = 0.008

    ws['o18'] = "2022.7~2022.12"
    ws['p18'] = 0.045
    ws['q18'] = 0.03495
    ws['r18'] = 0.1227
    ws['s18'] = 0.009
    
    filename = os.path.splitext(i)[0]
    extention = os.path.splitext(i)[1]

    wb.save(filename+'(py).xlsm')

    print(i+'파일을 완료했습니다')
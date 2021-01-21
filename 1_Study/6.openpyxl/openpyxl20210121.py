import openpyxl
import os
import pandas as pd
import shutil
import re
import openpyxl


# 대상파일 획정하기
path='D:\\NaverCloud\\화랑\\☆ 자문'
how='2021년\s*임금대장\s*.+[.]x+'

empty_list=[]

def checkfile(path, target):

    cp=re.compile(target)

    for i in os.listdir(path):
        abs_path = os.path.join(path,i)
        if os.path.isfile(abs_path):    #os.path.join 함수는 절대경로가 있는 파일만 Read 할 수 있는 것 같음
            if cp.match(i):
                empty_list.append(abs_path)
        
        elif os.path.isdir(abs_path):
            checkfile(abs_path, target)

checkfile(path,how)

# 대상파일을 현재작업폴더에 복사하기
for i in empty_list:
    wb = openpyxl.load_workbook(i)
    ws = wb['(조건)']
    ws['p6'] = 0.1152
    ws['r16'] = 0.1152
    wb.save(path)
    print(i+'파일을 완료했습니다')